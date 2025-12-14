"""
Utility functions for merging Excel files.
"""
import os
import openpyxl 
from openpyxl.styles import Alignment

def copy_data(source_file, target_file):
    """
    Copy data from source Excel file to target Excel file.
    
    Args:
        source_file: Path to source Excel file
        target_file: Path to target Excel file (will be created if not exists)
    """
    # Load source file
    source_wb = openpyxl.load_workbook(source_file)
    source_sheet = source_wb.active  # Get active worksheet from source file

    # Try to load target file, create new workbook if it doesn't exist
    try:
        target_wb = openpyxl.load_workbook(target_file)
        target_sheet = target_wb.active  # Get active worksheet from target file
    except FileNotFoundError:
        target_wb = openpyxl.Workbook()
        target_sheet = target_wb.active  # Create new worksheet

    # Iterate through rows and copy data
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)  # Append each row to target worksheet

    # Save target workbook
    target_wb.save(target_file)
    print(f'Data has been copied from {source_file} to {target_file}')
